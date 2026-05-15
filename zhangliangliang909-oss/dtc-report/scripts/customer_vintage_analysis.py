#!/usr/bin/env python3
"""
客户年代分析模块

功能：
1. 统计各年合作客户数（2022-2026）
2. 分析各年开发客户对本年收入的贡献
"""

import openpyxl
import glob
import os
from collections import defaultdict
from datetime import datetime, timedelta

# 数据目录
BUSINESS_DATA_DIR = "/home/admin/.openclaw/workspace/agents/跨境电商财务分析-agent/data/1.业务和订单数据"


def analyze_customer_vintage(period='2026-Q1'):
    """
    客户年代分析
    
    返回：
    - 各年合作客户数统计
    - 各年开发客户对本年收入的贡献
    """
    print('  进行客户年代分析...')
    
    # 1. 读取历史客户合作数据，确定每个客户的首次合作年份
    customer_first_year = read_customer_first_cooperation_year()
    
    # 2. 读取 2022-2026 年各年合作客户数
    yearly_customers = read_yearly_cooperation_customers(customer_first_year)
    
    # 3. 读取 2026 Q1 收入数据，按客户汇总
    current_revenue_by_customer = read_current_revenue_by_customer(period)
    
    # 4. 分析各年开发客户对本年收入的贡献
    revenue_contribution_by_vintage = analyze_revenue_contribution_by_vintage(
        customer_first_year, 
        current_revenue_by_customer
    )
    
    return {
        'yearly_customers': yearly_customers,
        'revenue_contribution_by_vintage': revenue_contribution_by_vintage,
        'total_unique_customers': len(customer_first_year)
    }


def read_customer_first_cooperation_year():
    """
    读取每个客户的首次合作年份
    
    从 DTC 历史客户合作情况汇总.xlsx 读取
    """
    customer_first_year = {}
    
    files = glob.glob(os.path.join(BUSINESS_DATA_DIR, '*历史客户合作情况汇总*.xlsx'))
    if not files:
        print('    ⚠️ 未找到历史客户合作情况汇总文件')
        return customer_first_year
    
    f = files[0]
    wb = openpyxl.load_workbook(f, data_only=True)
    ws = wb.active
    
    # 找表头
    headers = {}
    for col in range(1, 20):
        h = ws.cell(row=1, column=col).value
        if h:
            headers[str(h).strip()] = col
    
    customer_col = headers.get('客户名称', 1)
    year_col = headers.get('第一次合作年', 4)
    segment_col = headers.get('业务分段列表', 2)
    
    for row in range(2, min(ws.max_row + 1, 10000)):
        customer = ws.cell(row=row, column=customer_col).value
        first_year = ws.cell(row=row, column=year_col).value
        segment = ws.cell(row=row, column=segment_col).value
        
        if not customer or not first_year:
            continue
        
        # 处理业务分段
        segments = []
        seg_str = str(segment).strip() if segment else ''
        if 'A' in seg_str or 'B' in seg_str:
            segments.append('b')
        if 'C' in seg_str:
            segments.append('c')
        if 'D' in seg_str or '电商' in seg_str or '集拼' in seg_str:
            segments.append('d')
        
        # 记录客户首次合作年份（分业务段）
        if customer not in customer_first_year:
            customer_first_year[customer] = {'b': None, 'c': None, 'd': None}
        
        for seg in segments:
            try:
                year = int(first_year)
                if customer_first_year[customer][seg] is None or year < customer_first_year[customer][seg]:
                    customer_first_year[customer][seg] = year
            except:
                continue
    
    print(f'    读取历史客户：{len(customer_first_year)}家')
    return customer_first_year


def read_yearly_cooperation_customers(customer_first_year):
    """
    统计 2022-2026 年各年合作客户数
    
    有正向收入即视为合作
    """
    yearly_customers = {
        '2022': {'b': set(), 'c': set(), 'd': set()},
        '2023': {'b': set(), 'c': set(), 'd': set()},
        '2024': {'b': set(), 'c': set(), 'd': set()},
        '2025': {'b': set(), 'c': set(), 'd': set()},
        '2026': {'b': set(), 'c': set(), 'd': set()}
    }
    
    files = glob.glob(os.path.join(BUSINESS_DATA_DIR, '*所有业务收入明细*.xlsx'))
    if not files:
        return yearly_customers
    
    f = files[0]
    wb = openpyxl.load_workbook(f, data_only=True)
    ws = wb.active
    
    # 找表头
    headers = {}
    for col in range(1, 20):
        h = ws.cell(row=1, column=col).value
        if h:
            headers[str(h).strip()] = col
    
    bu_col = headers.get('业务系统单元编码', 1)
    segment_col = headers.get('业务分段分类_新.', 2)
    month_col = headers.get('业务年月', 3)
    customer_col = headers.get('委托客户名称', 4)
    revenue_col = headers.get('收入', 8)
    
    for row in range(2, min(ws.max_row + 1, 50000)):
        bu = ws.cell(row=row, column=bu_col).value
        segment = ws.cell(row=row, column=segment_col).value
        month = ws.cell(row=row, column=month_col).value
        customer = ws.cell(row=row, column=customer_col).value
        revenue = ws.cell(row=row, column=revenue_col).value
        
        if not bu or str(bu).strip() != 'BWLDTC':
            continue
        
        if not customer or not revenue or revenue <= 0:
            continue
        
        # 判断业务段
        seg_str = str(segment).strip() if segment else ''
        segments = []
        if 'B 段' in seg_str or 'B' in seg_str or 'A+B' in seg_str:
            segments.append('b')
        if 'C 段' in seg_str or 'C' in seg_str or 'C+D' in seg_str:
            segments.append('c')
        if 'D 段' in seg_str or 'D' in seg_str or '电商' in seg_str or '集拼' in seg_str:
            segments.append('d')
        
        # 判断年份
        year = None
        if isinstance(month, (int, float)):
            base_date = datetime(1899, 12, 30)
            actual_date = base_date + timedelta(days=int(month))
            year = str(actual_date.year)
        elif isinstance(month, str):
            if '年' in month:
                year = month.split('年')[0].strip()
            elif '-' in month:
                year = month.split('-')[0].strip()
        
        if year not in ['2022', '2023', '2024', '2025', '2026']:
            continue
        
        # 记录合作客户
        for seg in segments:
            yearly_customers[year][seg].add(customer)
    
    # 统计各年客户数
    result = {}
    for year in ['2022', '2023', '2024', '2025', '2026']:
        result[year] = {
            'b': len(yearly_customers[year]['b']),
            'c': len(yearly_customers[year]['c']),
            'd': len(yearly_customers[year]['d']),
            'total': len(yearly_customers[year]['b'] | yearly_customers[year]['c'] | yearly_customers[year]['d'])
        }
    
    return result


def read_current_revenue_by_customer(period='2026-Q1'):
    """
    读取 2026 Q1 各客户收入（分业务段）
    """
    current_month, cumulative_months = parse_period(period)
    
    revenue_by_customer = defaultdict(lambda: {'b': 0, 'c': 0, 'd': 0})
    
    files = glob.glob(os.path.join(BUSINESS_DATA_DIR, '*所有业务收入明细*.xlsx'))
    if not files:
        return revenue_by_customer
    
    f = files[0]
    wb = openpyxl.load_workbook(f, data_only=True)
    ws = wb.active
    
    # 找表头
    headers = {}
    for col in range(1, 20):
        h = ws.cell(row=1, column=col).value
        if h:
            headers[str(h).strip()] = col
    
    bu_col = headers.get('业务系统单元编码', 1)
    segment_col = headers.get('业务分段分类_新.', 2)
    month_col = headers.get('业务年月', 3)
    customer_col = headers.get('委托客户名称', 4)
    revenue_col = headers.get('收入', 8)
    
    for row in range(2, min(ws.max_row + 1, 50000)):
        bu = ws.cell(row=row, column=bu_col).value
        segment = ws.cell(row=row, column=segment_col).value
        month = ws.cell(row=row, column=month_col).value
        customer = ws.cell(row=row, column=customer_col).value
        revenue = ws.cell(row=row, column=revenue_col).value
        
        if not bu or str(bu).strip() != 'BWLDTC':
            continue
        
        if not customer or not revenue:
            continue
        
        # 判断业务段
        seg_str = str(segment).strip() if segment else ''
        segments = []
        if 'B 段' in seg_str or 'B' in seg_str or 'A+B' in seg_str:
            segments.append('b')
        if 'C 段' in seg_str or 'C' in seg_str or 'C+D' in seg_str:
            segments.append('c')
        if 'D 段' in seg_str or 'D' in seg_str or '电商' in seg_str or '集拼' in seg_str:
            segments.append('d')
        
        # 判断月份
        month_num = None
        if isinstance(month, (int, float)):
            base_date = datetime(1899, 12, 30)
            actual_date = base_date + timedelta(days=int(month))
            month_num = actual_date.month
        elif isinstance(month, str) and '-' in month:
            try:
                month_num = int(month.split('-')[1])
            except:
                continue
        
        if month_num not in cumulative_months:
            continue
        
        # 累加收入（转换为万元）
        for seg in segments:
            revenue_by_customer[customer][seg] += float(revenue) / 10000
    
    return revenue_by_customer


def analyze_revenue_contribution_by_vintage(customer_first_year, current_revenue_by_customer):
    """
    分析各年开发客户对本年收入的贡献
    
    按客户首次合作年份分组，统计 2026 Q1 收入贡献
    """
    contribution = {
        '2022': {'b': 0, 'c': 0, 'd': 0, 'customers': set()},
        '2023': {'b': 0, 'c': 0, 'd': 0, 'customers': set()},
        '2024': {'b': 0, 'c': 0, 'd': 0, 'customers': set()},
        '2025': {'b': 0, 'c': 0, 'd': 0, 'customers': set()},
        '2026': {'b': 0, 'c': 0, 'd': 0, 'customers': set()},
        'unknown': {'b': 0, 'c': 0, 'd': 0, 'customers': set()}
    }
    
    for customer, revenue in current_revenue_by_customer.items():
        # 获取客户首次合作年份
        first_year = customer_first_year.get(customer, {})
        
        for seg in ['b', 'c', 'd']:
            if revenue[seg] > 0:
                # 确定年份
                year = first_year.get(seg, None)
                if year and str(year) in ['2022', '2023', '2024', '2025', '2026']:
                    vintage = str(year)
                else:
                    vintage = 'unknown'
                
                contribution[vintage][seg] += revenue[seg]
                contribution[vintage]['customers'].add(customer)
    
    # 格式化结果
    result = {}
    for vintage in ['2022', '2023', '2024', '2025', '2026', 'unknown']:
        result[vintage] = {
            'b_revenue': contribution[vintage]['b'],
            'c_revenue': contribution[vintage]['c'],
            'd_revenue': contribution[vintage]['d'],
            'total_revenue': contribution[vintage]['b'] + contribution[vintage]['c'] + contribution[vintage]['d'],
            'customer_count': len(contribution[vintage]['customers'])
        }
    
    return result


def parse_period(period):
    """解析期间参数"""
    if period.upper().startswith('2026-Q'):
        quarter = int(period[-1])
        current_month = quarter * 3
        cumulative_months = list(range(1, current_month + 1))
        return current_month, cumulative_months
    
    if period.startswith('2026-'):
        month_part = period[5:]
        current_month = int(month_part)
        cumulative_months = list(range(1, current_month + 1))
        return current_month, cumulative_months
    
    return 3, [1, 2, 3]


def format_customer_vintage_html(analysis_result):
    """将客户年代分析结果格式化为 HTML"""
    if not analysis_result:
        return ''
    
    html = '''
    <div class="section">
        <div class="section-title">四、客户与销售分析</div>
        
        <div class="subsection">
            <div class="subsection-title">4.1 客户开发数</div>
            <table>
                <thead>
                    <tr>
                        <th>年份</th>
                        <th>B 段客户数</th>
                        <th>C 段客户数</th>
                        <th>D 段客户数</th>
                        <th>合计（去重）</th>
                    </tr>
                </thead>
                <tbody>
'''
    
    yearly = analysis_result.get('yearly_customers', {})
    for year in ['2022', '2023', '2024', '2025', '2026']:
        data = yearly.get(year, {})
        html += f'''
                    <tr>
                        <td>{year}年</td>
                        <td>{data.get('b', 0):,}</td>
                        <td>{data.get('c', 0):,}</td>
                        <td>{data.get('d', 0):,}</td>
                        <td>{data.get('total', 0):,}</td>
                    </tr>
'''
    
    html += f'''
                    <tr style="background: #f3f4f6; font-weight: 600;">
                        <td>合计</td>
                        <td colspan="4">累计合作客户总数：{analysis_result.get('total_unique_customers', 0):,}家（去重）</td>
                    </tr>
                </tbody>
            </table>
        </div>
        
        <div class="subsection">
            <div class="subsection-title">4.2 各年开发客户对本年收入的贡献（2026 Q1）</div>
            <table>
                <thead>
                    <tr>
                        <th>首次合作年份</th>
                        <th>B 段收入（万元）</th>
                        <th>C 段收入（万元）</th>
                        <th>D 段收入（万元）</th>
                        <th>合计（万元）</th>
                        <th>客户数</th>
                        <th>占比</th>
                    </tr>
                </thead>
                <tbody>
'''
    
    vintage_data = analysis_result.get('revenue_contribution_by_vintage', {})
    total_revenue = sum(vintage_data.get(v, {}).get('total_revenue', 0) for v in ['2022', '2023', '2024', '2025', '2026'])
    
    for vintage in ['2022', '2023', '2024', '2025', '2026']:
        data = vintage_data.get(vintage, {})
        rev = data.get('total_revenue', 0)
        percentage = (rev / total_revenue * 100) if total_revenue > 0 else 0
        html += f'''
                    <tr>
                        <td>{vintage}年</td>
                        <td>{data.get('b_revenue', 0):,.1f}</td>
                        <td>{data.get('c_revenue', 0):,.1f}</td>
                        <td>{data.get('d_revenue', 0):,.1f}</td>
                        <td>{rev:,.1f}</td>
                        <td>{data.get('customer_count', 0):,}</td>
                        <td>{percentage:.1f}%</td>
                    </tr>
'''
    
    # 未知年份
    unknown_data = vintage_data.get('unknown', {})
    if unknown_data.get('total_revenue', 0) > 0:
        rev = unknown_data.get('total_revenue', 0)
        percentage = (rev / total_revenue * 100) if total_revenue > 0 else 0
        html += f'''
                    <tr>
                        <td>未知</td>
                        <td>{unknown_data.get('b_revenue', 0):,.1f}</td>
                        <td>{unknown_data.get('c_revenue', 0):,.1f}</td>
                        <td>{unknown_data.get('d_revenue', 0):,.1f}</td>
                        <td>{rev:,.1f}</td>
                        <td>{unknown_data.get('customer_count', 0):,}</td>
                        <td>{percentage:.1f}%</td>
                    </tr>
'''
    
    html += f'''
                    <tr style="background: #f3f4f6; font-weight: 600;">
                        <td>合计</td>
                        <td>{sum(vintage_data.get(v, {}).get('b_revenue', 0) for v in ['2022', '2023', '2024', '2025', '2026']):,.1f}</td>
                        <td>{sum(vintage_data.get(v, {}).get('c_revenue', 0) for v in ['2022', '2023', '2024', '2025', '2026']):,.1f}</td>
                        <td>{sum(vintage_data.get(v, {}).get('d_revenue', 0) for v in ['2022', '2023', '2024', '2025', '2026']):,.1f}</td>
                        <td>{total_revenue:,.1f}</td>
                        <td>{sum(vintage_data.get(v, {}).get('customer_count', 0) for v in ['2022', '2023', '2024', '2025', '2026']):,}</td>
                        <td>100.0%</td>
                    </tr>
                </tbody>
            </table>
            <div class="data-source" style="margin-top: 10px;">
                注：按客户首次合作年份分组，统计 2026 Q1 收入贡献；合计客户数可能大于实际客户数（因同一客户可能合作多个业务段）
            </div>
        </div>
    </div>
'''
    
    return html


if __name__ == '__main__':
    print('客户年代分析模块加载成功')
